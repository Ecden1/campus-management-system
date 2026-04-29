from flask import render_template, url_for, flash, redirect, request, send_file
from flask_login import login_required, current_user
from app import app, db
from app.models import Student, Class, StudentStatus, OperationLog, Teacher, StudentCourse
from app.views.student_forms import StudentForm, StudentUpdateForm
import openpyxl
import io
from datetime import datetime

def log_operation(user_id, operation_type, operation_content, operation_ip, operation_result):
    log = OperationLog(
        user_id=user_id,
        operation_type=operation_type,
        operation_content=operation_content,
        operation_ip=operation_ip,
        operation_result=operation_result
    )
    db.session.add(log)
    db.session.commit()


# 显示所有学生的列表。如果当前用户是教师，那么他们只能看到选择他们教授的课程的学生。
@login_required
def index():
    if current_user.role == 'teacher':#检查身份信息
        # 教师只能查看选择自己教授课程的学生
        teacher = Teacher.query.filter_by(teacher_id=current_user.username).first()
        if teacher:
            # 获取教师教授的课程
            course_id = teacher.course_id
            # 获取选择该课程的学生
            student_courses = StudentCourse.query.filter_by(course_id=course_id).all()
            student_ids = [sc.student_id for sc in student_courses]
            students = Student.query.filter(Student.student_id.in_(student_ids)).all()
        else:
            students = []
    else:
        students = Student.query.all()
    return render_template('student/index.html', students=students, title='学生管理')


# 添加新学生的视图函数
@login_required
def add():
    if current_user.role == 'student':
        flash('权限不足', 'danger')
        return redirect(url_for('student.index'))
    form = StudentForm()
    form.class_id.choices = [(c.id, c.class_name) for c in Class.query.all()]
    if form.validate_on_submit():
        student = Student(
            student_id=form.student_id.data,
            name=form.name.data,
            gender=form.gender.data,
            birthday=form.birthday.data,
            id_card=form.id_card.data,
            phone=form.phone.data,
            address=form.address.data,
            class_id=form.class_id.data
        )
        db.session.add(student)
        # 创建初始学籍状态
        status = StudentStatus(
            student_id=student.student_id,
            status='在读',
            operator_id=current_user.id
        )
        db.session.add(status)
        db.session.commit()
        log_operation(current_user.id, 'add_student', f'添加学生: {student.student_id} - {student.name}', request.remote_addr, 'success')
        flash('学生添加成功', 'success')
        return redirect(url_for('student.index'))
    return render_template('student/add.html', form=form, title='添加学生')


# 更新学生信息的视图函数。
@login_required
def update(id):
    if current_user.role == 'student':
        flash('权限不足', 'danger')
        return redirect(url_for('student.index'))
    student = Student.query.get_or_404(id)
    form = StudentUpdateForm(obj=student)
    form.class_id.choices = [(c.id, c.class_name) for c in Class.query.all()]
    if form.validate_on_submit():
        student.student_id = form.student_id.data
        student.name = form.name.data
        student.gender = form.gender.data
        student.birthday = form.birthday.data
        student.id_card = form.id_card.data
        student.phone = form.phone.data
        student.address = form.address.data
        student.class_id = form.class_id.data
        db.session.commit()
        log_operation(current_user.id, 'update_student', f'更新学生: {student.student_id} - {student.name}', request.remote_addr, 'success')
        flash('学生信息更新成功', 'success')
        return redirect(url_for('student.index'))
    return render_template('student/update.html', form=form, student=student, title='更新学生')

@login_required
def delete(id):
    if current_user.role == 'student':
        flash('权限不足', 'danger')
        return redirect(url_for('student.index'))
    student = Student.query.get_or_404(id)
    db.session.delete(student)
    db.session.commit()
    log_operation(current_user.id, 'delete_student', f'删除学生: {student.student_id} - {student.name}', request.remote_addr, 'success')
    flash('学生删除成功', 'success')
    return redirect(url_for('student.index'))



# 从Excel文件导入学生数据的视图函数
@login_required
def import_students():
    if current_user.role == 'student':
        flash('权限不足', 'danger')
        return redirect(url_for('student.index'))
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择文件', 'danger')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('请选择文件', 'danger')
            return redirect(request.url)
        if file and file.filename.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2):
                    student_id = str(row[0].value)
                    name = row[1].value
                    gender = row[2].value
                    birthday = row[3].value
                    id_card = str(row[4].value)
                    phone = str(row[5].value)
                    address = row[6].value
                    class_name = row[7].value
                    
                    class_ = Class.query.filter_by(class_name=class_name).first()
                    if not class_:
                        continue
                    
                    student = Student(
                        student_id=student_id,
                        name=name,
                        gender=gender,
                        birthday=birthday,
                        id_card=id_card,
                        phone=phone,
                        address=address,
                        class_id=class_.id
                    )
                    db.session.add(student)
                    
                    status = StudentStatus(
                        student_id=student_id,
                        status='在读',
                        operator_id=current_user.id
                    )
                    db.session.add(status)
                db.session.commit()
                log_operation(current_user.id, 'import_students', '批量导入学生', request.remote_addr, 'success')
                flash('学生批量导入成功', 'success')
            except Exception as e:
                log_operation(current_user.id, 'import_students', f'批量导入学生失败: {str(e)}', request.remote_addr, 'failed')
                flash(f'导入失败: {str(e)}', 'danger')
            return redirect(url_for('student.index'))
    return render_template('student/import.html', title='批量导入学生')


# 将所有学生的数据导出到Excel文件的视图函数
@login_required
def export_students():
    if current_user.role == 'student':
        flash('权限不足', 'danger')
        return redirect(url_for('student.index'))
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '学生信息'
    
    headers = ['学号', '姓名', '性别', '出生日期', '身份证号', '手机号', '地址', '班级']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    students = Student.query.all()
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.student_id)
        ws.cell(row=row, column=2, value=student.name)
        ws.cell(row=row, column=3, value=student.gender)
        ws.cell(row=row, column=4, value=student.birthday.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=student.id_card)
        ws.cell(row=row, column=6, value=student.phone)
        ws.cell(row=row, column=7, value=student.address)
        ws.cell(row=row, column=8, value=student.class_.class_name)
    
    wb.save(output)
    output.seek(0)
    log_operation(current_user.id, 'export_students', '导出学生信息', request.remote_addr, 'success')
    return send_file(output, as_attachment=True, download_name=f'students_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

from flask_wtf import FlaskForm
from wtforms import StringField, DateField, TextAreaField
from wtforms.validators import DataRequired, Length, Regexp

class EditProfileForm(FlaskForm):
    birthday = DateField('出生日期', validators=[DataRequired()])
    id_card = StringField('身份证号', validators=[DataRequired(), Length(min=18, max=18)])
    phone = StringField('手机号', validators=[DataRequired(), Length(min=11, max=11), Regexp('^1[3-9]\d{9}$')])
    address = TextAreaField('地址', validators=[DataRequired(), Length(max=200)])


# 显示当前登录学生的详细信息的视图函数
@login_required
def profile():
    # 假设学生的用户名就是学号
    student = Student.query.filter_by(student_id=current_user.username).first()
    return render_template('student/profile.html', student=student, editing=False, title='个人信息')


# 更新当前登录学生的详细信息的视图函数。
@login_required
def edit_profile():
    student = Student.query.filter_by(student_id=current_user.username).first()
    if not student:
        flash('未找到个人信息', 'danger')
        return redirect(url_for('student.profile'))
    
    form = EditProfileForm(obj=student)
    if form.validate_on_submit():
        student.birthday = form.birthday.data
        student.id_card = form.id_card.data
        student.phone = form.phone.data
        student.address = form.address.data
        db.session.commit()
        log_operation(current_user.id, 'edit_profile', f'修改个人信息: {student.student_id} - {student.name}', request.remote_addr, 'success')
        flash('个人信息更新成功', 'success')
        return redirect(url_for('student.profile'))
    return render_template('student/profile.html', student=student, form=form, editing=True, title='编辑个人信息')