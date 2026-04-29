from flask import render_template, url_for, request, send_file, flash, redirect
from flask_login import login_required, current_user
from app import app, db
from app.models import Student, Class, Course, Grade
import openpyxl
import io
from datetime import datetime


# 该 report 模块提供了关于班级信息和成绩的报表分析功能


# index()：返回报表分析首页
def index():
    return render_template('report/index.html', title='报表分析')

# class_report()：获取所有班级并展示班级报表页面。
def class_report():
    classes = Class.query.all()
    return render_template('report/class_report.html', classes=classes, title='班级报表')

# 获取所有课程并展示成绩报表页面。
def grade_report():
    courses = Course.query.all()
    return render_template('report/grade_report.html', courses=courses, title='成绩报表')


# 导出指定班级的学生信息报表。首先获取指定的班级以及该班级的所有学生，
# 然后创建一个 Excel 工作簿，将学生的所有信息写入工作簿，并返回此工作簿作为文件下载
def export_class_report(class_id):
    if current_user.role == 'student':
        flash('权限不足', 'danger')
        return redirect(url_for('report.class_report'))
    
    class_ = Class.query.get_or_404(class_id)
    students = class_.students
    
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{class_.class_name} 学生信息'
    
    headers = ['学号', '姓名', '性别', '出生日期', '身份证号', '手机号', '地址', '学籍状态']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.student_id)
        ws.cell(row=row, column=2, value=student.name)
        ws.cell(row=row, column=3, value=student.gender)
        ws.cell(row=row, column=4, value=student.birthday.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=student.id_card)
        ws.cell(row=row, column=6, value=student.phone)
        ws.cell(row=row, column=7, value=student.address)
        ws.cell(row=row, column=8, value=student.status.status if student.status else '未知')
    
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'{class_.class_name}_students_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# 导出指定课程的成绩报表。
def export_grade_report(course_id):
    course = Course.query.get_or_404(course_id)
    
    if current_user.role == 'student':
        # 学生只能导出自己的成绩
        grades = Grade.query.filter_by(student_id=current_user.username, course_id=course_id).all()
    else:
        grades = course.grades
    
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{course.course_name} 成绩'
    
    headers = ['学号', '姓名', '平时成绩', '期末成绩', '总评成绩', '状态']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row, grade in enumerate(grades, 2):
        student = Student.query.filter_by(student_id=grade.student_id).first()
        ws.cell(row=row, column=1, value=grade.student_id)
        ws.cell(row=row, column=2, value=student.name if student else '')
        ws.cell(row=row, column=3, value=grade.daily_score)
        ws.cell(row=row, column=4, value=grade.exam_score)
        ws.cell(row=row, column=5, value=grade.total_score)
        ws.cell(row=row, column=6, value='已锁定' if grade.status == 'locked' else '未锁定')
    
    # 计算统计信息
    if grades:
        total_scores = [g.total_score for g in grades]
        avg_score = sum(total_scores) / len(total_scores)
        max_score = max(total_scores)
        min_score = min(total_scores)
        pass_count = sum(1 for g in grades if g.total_score >= 60)
        pass_rate = pass_count / len(grades) * 100
        
        ws.cell(row=len(grades) + 3, column=1, value='统计信息')
        ws.cell(row=len(grades) + 4, column=1, value='平均分')
        ws.cell(row=len(grades) + 4, column=2, value=avg_score)
        ws.cell(row=len(grades) + 5, column=1, value='最高分')
        ws.cell(row=len(grades) + 5, column=2, value=max_score)
        ws.cell(row=len(grades) + 6, column=1, value='最低分')
        ws.cell(row=len(grades) + 6, column=2, value=min_score)
        ws.cell(row=len(grades) + 7, column=1, value='及格率')
        ws.cell(row=len(grades) + 7, column=2, value=f'{pass_rate:.2f}%')
    
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'{course.course_name}_grades_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# 导出当前登录学生的所有成绩
def export_student_grades():
    if current_user.role != 'student':
        flash('权限不足', 'danger')
        return redirect(url_for('report.grade_report'))
    
    # 获取学生的所有成绩
    grades = Grade.query.filter_by(student_id=current_user.username).all()
    
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '个人所有成绩'
    
    headers = ['课程代码', '课程名称', '学分', '平时成绩', '期末成绩', '总评成绩', '状态']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row, grade in enumerate(grades, 2):
        course = Course.query.get(grade.course_id)
        ws.cell(row=row, column=1, value=course.course_code if course else '')
        ws.cell(row=row, column=2, value=course.course_name if course else '')
        ws.cell(row=row, column=3, value=course.credit if course else '')
        ws.cell(row=row, column=4, value=grade.daily_score)
        ws.cell(row=row, column=5, value=grade.exam_score)
        ws.cell(row=row, column=6, value=grade.total_score)
        ws.cell(row=row, column=7, value='已锁定' if grade.status == 'locked' else '未锁定')
    
    # 计算统计信息
    if grades:
        total_scores = [g.total_score for g in grades]
        avg_score = sum(total_scores) / len(total_scores)
        max_score = max(total_scores)
        min_score = min(total_scores)
        pass_count = sum(1 for g in grades if g.total_score >= 60)
        pass_rate = pass_count / len(grades) * 100
        total_credits = sum(Course.query.get(g.course_id).credit for g in grades if Course.query.get(g.course_id))
        
        ws.cell(row=len(grades) + 3, column=1, value='统计信息')
        ws.cell(row=len(grades) + 4, column=1, value='课程总数')
        ws.cell(row=len(grades) + 4, column=2, value=len(grades))
        ws.cell(row=len(grades) + 5, column=1, value='总学分')
        ws.cell(row=len(grades) + 5, column=2, value=total_credits)
        ws.cell(row=len(grades) + 6, column=1, value='平均分')
        ws.cell(row=len(grades) + 6, column=2, value=avg_score)
        ws.cell(row=len(grades) + 7, column=1, value='最高分')
        ws.cell(row=len(grades) + 7, column=2, value=max_score)
        ws.cell(row=len(grades) + 8, column=1, value='最低分')
        ws.cell(row=len(grades) + 8, column=2, value=min_score)
        ws.cell(row=len(grades) + 9, column=1, value='及格率')
        ws.cell(row=len(grades) + 9, column=2, value=f'{pass_rate:.2f}%')
    
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'个人所有成绩_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')